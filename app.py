'''
    Created by Ximena Leyva Peralta for Yale's CS50 Fall 2022 class

        References:
        - https://www.youtube.com/watch?v=Rq2G_jfL_4g&ab_channel=TechnoDine
        - https://github.com/realpython/flask-bokeh-example/blob/master/tutorial.md
        - https://www.geeksforgeeks.org/how-to-iterate-through-excel-rows-in-python/ 
        - https://openpyxl.readthedocs.io/en/latest/api/openpyxl.workbook.workbook.html
        - https://docs.bokeh.org/en/latest/
        - https://tedboy.github.io/flask/generated/flask.send_file.html
        - https://www.sqlite.org/docs.html
        - https://flask.palletsprojects.com/en/2.2.x/patterns/flashing/
        - https://sebhastian.com/javascript-confirmation-yes-no/#:~:text=You%20can%20create%20a%20JavaScript,can%20specify%20as%20its%20argument.
        - https://getbootstrap.com/docs/5.2/getting-started/introduction/
'''



''' LIBRARIES '''

# handle sql database
from cs50 import SQL

# create flask app
from flask import Flask, flash, redirect, render_template, request, session, send_file
from flask_session import Session

# utilities
from functools import wraps
from werkzeug.security import check_password_hash, generate_password_hash

# work with Excel
from openpyxl import Workbook, load_workbook

# make interactive plots
from bokeh.embed import components
from bokeh.plotting import figure
from bokeh.models import FactorRange

# format dates
import calendar





''' SET UP '''

# create application
app = Flask(__name__)

# configure sessions for the app
app.config["SESSION_PERMANENT"] = False
app.config["SESSION_TYPE"] = "filesystem"
Session(app)

# set up database
db = SQL("sqlite:///BF_admin_site.db")

# path with Excel file
path = './finances_BF.xlsx'


# functions to improve app
# from CS50 Finance Pset
@app.after_request
def after_request(response):
    """Ensure responses aren't cached"""
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Expires"] = 0
    response.headers["Pragma"] = "no-cache"
    return response

def login_required(f):
    """
    Decorate routes to require login.

    https://flask.palletsprojects.com/en/1.1.x/patterns/viewdecorators/
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("user_id") is None:
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated_function





''' MAIN PAGE '''
@app.route("/")
@login_required
def index():
    ''' Renders home page and welcomes the user '''
    
    # get the current user's name from the members database
    user_id = session["user_id"]
    member_id = db.execute('SELECT member_id FROM users WHERE id = ?', user_id)[0]['member_id']
    name = db.execute('SELECT name FROM members WHERE id = ?', member_id)[0]['name']
    
    # render template of home page
    return render_template('index.html', name=name)






''' HANDLING ACCOUNTS '''

@app.route("/login", methods=["GET", "POST"])
def login():
    """ Log user in """
    
    # forget any user_id
    session.clear()
    
    # user reached via POST
    if request.method == "POST":

        # ensure username was submitted
        if not request.form.get("username"):
            flash('Type a username')
            return redirect('/login')

        # ensure password was submitted
        elif not request.form.get("password"):
            flash('Type a password')
            return redirect('/login')

        # query database for username & password
        rows = db.execute("SELECT * FROM users WHERE username = ?", request.form.get("username"))
        pw = request.form.get("password")
        
        # ensure username exists and password is correct
        if len(rows) != 1 or not check_password_hash(rows[0]["hash"], pw):  # type: ignore
            flash("Incorrect username or password")
            return redirect('/login')

        # remember which user has logged in
        session["user_id"] = rows[0]["id"]
        
        # redirect user to home page
        return redirect("/")

    # Reached by GET only renders template
    else:
        return render_template('login.html')


@app.route("/logout")
def logout():
    """ Log user out """

    # forget any user_id
    session.clear()

    # redirect user to login form
    return redirect("/")


@app.route("/register", methods=["GET", "POST"])
def register():
    """ Register user """
    
    if request.method == "POST":
        
        # get username, email and the id of the member
        username = request.form.get("username")
        email = request.form.get("email")
        member_id = request.form.get("member")
        
        # ensure username was submitted
        if not username:
            flash('Type a username')
            return redirect('/register')
        
        # username cannot contain spaces
        elif ' ' in username:  # type: ignore
            flash("Username cannot contain spaces")
            return redirect('/register')
        
        # ensure email was submitted
        elif not email:
            flash('Type an email')
            return redirect('/register')  # in future extensions of the program, email could be used for password recovery

        # ensure password was submitted
        elif not request.form.get("password"):
            flash('Type a password')
            return redirect('/register')

        # ensure password confirmation was submitted
        elif not request.form.get("confirmation"):
            flash('Retype password')
            return redirect('/register')

        # ensure password and confirmation match
        elif request.form.get("confirmation") != request.form.get("password"):
            flash("Passwords don't match")
            return redirect('/register')
        
        # ensure a member was selected from dropdown
        elif not member_id:
            flash("Choose your name from the members list")
            return redirect('/register')

        # make sure username does not already exist
        look_username = db.execute("SELECT username FROM users WHERE username = ?", username)
        if look_username:
            flash("Username already exists")
            return redirect('/register')

        # hash the password
        hashed = generate_password_hash(request.form.get("password"))  # type: ignore
        
        # input new row into users dataset
        db.execute("INSERT INTO users (username, hash, email, member_id) VALUES (?, ?, ?, ?)", username, hashed, email, member_id)

        # return to log in page
        return redirect("/")
    
    else:
        # get the list of group members to be used in dropdown
        members = db.execute("SELECT name, id FROM members WHERE id NOT IN (SELECT member_id FROM members JOIN users ON members.id = users.member_id);")
       
        return render_template("register.html", members=members)
    

@app.route("/site_accounts")
@login_required
def site_accounts():
    ''' Access to accounts registered in app '''
    
    # in future versions of the program, this section would only be accessible to the Co-Presidents and Treasurer of the dance group
    
    # get the list of accounts in the database
    accounts = db.execute("SELECT username, name FROM users JOIN members ON users.member_id = members.id")
    
    # get username of active user
    user_id = session["user_id"]
    user_name = db.execute("SELECT username FROM users WHERE id = ?", user_id)[0]['username']
    
    for account in accounts:
        
        # do not display the account of the active user to prevent them from deleting it 
        if account['username'] == user_name:
            accounts.remove(account)
        
    # pass the account list for template to loop through and display
    return render_template('site_accounts.html', accounts=accounts)


@app.route("/delete_account", methods=["GET", "POST"])
@login_required
def delete_account():
    ''' Deletes an account '''
    
    # accessed through POST
    if request.method == 'POST':
        # get the id of the account to be deleted (from small form in html)
        account_id = request.form.get('id')
        
        # delete from database
        db.execute("DELETE FROM users WHERE id = ?", account_id)
    
    # return to acounts list
    return redirect('/site_accounts')






''' HANDLING MEMBERS & BOARD '''

@app.route("/members")
@login_required
def members():
    ''' Display list of group members '''
    
    # get all the members from database
    members = db.execute("SELECT * FROM members")

    # format if a member is active in the dance group
    for member in members:
        if member["active"] == 'True':
            member["active"] = 'Yes'
        elif member["active"] == 'False':
            member["active"] = 'No'
    
    # ensure active user is not shown their own group member profile to avoid them from deleting it
    user_id = session["user_id"]
    member_id = db.execute('SELECT member_id FROM users WHERE id = ?', user_id)[0]['member_id']
    
    for member in members:
        if member_id and member['id'] == member_id:
            members.remove(member)
    
    # render template with members list
    return render_template("member_list.html", members=members)

@app.route("/add_member", methods=["GET", "POST"])
@login_required
def add_member():
    ''' Add a new member to the database '''
    
    if request.method == "POST":
        
        # ensure a name was submitted
        if not request.form.get("name"):
            flash('Type name')
            return redirect('/add_member')
        
        # ensure a college was submitted
        if not request.form.get("college"):
            flash('Choose a college option')
            return redirect('/add_member')
        
        # get info from form
        name = request.form.get("name")
        college = request.form.get("college")
        year = request.form.get("year")
        active = request.form.get("active")
        
        # insert new row into database
        db.execute("INSERT INTO members (name, college, grad_year, active, board) VALUES (?, ?, ?, ?, 0)", name, college, year, active)
        
        # return to members list
        return redirect("/members")
    
    else:
        # format list of yale colleges
        colleges = ['JE', 'BR', 'SY', 'PC', 'DC', 'GH', 'BK', 'TC', 'ES', 'MC', 'TD', 'SM', 'BF', 'MY']
        
        # render template
        return render_template('add_member.html', colleges=colleges)


@app.route("/edit_member", methods=["GET", "POST"])
@login_required
def edit_member():
    ''' Loads page to edit group member '''
    
    if request.method == "POST":
        
        # get the id of the member to be editted
        member_id = request.form.get("member_id")
        
        # get member's info to pass in as current values in html form
        member_info = db.execute("SELECT * FROM members WHERE id = ?", member_id)[0]
        
        # render template with appropriate data 
        colleges = ['JE', 'BR', 'SY', 'PC', 'DC', 'GH', 'BK', 'TC', 'ES', 'MC', 'TD', 'SM', 'BF', 'MY']
        return render_template('edit_member.html', member=member_info, colleges=colleges)
   
    else:
        return redirect('/members')


@app.route("/editing_member", methods=["GET", "POST"])
@login_required
def editing_member():
    ''' Finalizes editing a group member '''
    
    if request.method == 'POST':
        
        # get the member who was editted
        member_id = request.form.get("member_id")
        
        # ensure a name was typed
        if not request.form.get('name'):
            flash('Type name')
            
            # return to edit_member interface with appropriate data
            member_info = db.execute("SELECT * FROM members WHERE id = ?", member_id)[0]
            colleges = ['JE', 'BR', 'SY', 'PC', 'DC', 'GH', 'BK', 'TC', 'ES', 'MC', 'TD', 'SM', 'BF', 'MY']
            return render_template('edit_member.html', member=member_info, colleges=colleges)

        # get info from form
        name = request.form.get("name")
        college = request.form.get("college")
        year = request.form.get("year")
        active = request.form.get("active")
        
        # update appropriate row in database
        db.execute("UPDATE members SET name = ?, college = ?, year = ?, active = ? WHERE id = ?", name, college, year, active, member_id)        
    
    # return to members page
    return redirect('/members')


@app.route("/delete_member", methods=["GET", "POST"])
@login_required
def delete_member():
    ''' Delete a group member '''
    
    if request.method == 'POST':
        
        # get the id of the member
        member_id = request.form.get("member_id")
        
        # remove them from the board
        db.execute("DELETE FROM board WHERE member_id = ?", member_id)
        
        # get the member_id of the current treasurer
        treasurer_id = db.execute("SELECT member_id FROM members JOIN board ON members.id = board.member_id WHERE position = 'Treasurer'")[0]['member_id']

        # all tables where the id of the deleted member was used are replaced with the treasurer's id
        # i.e. in all purchases or earnings the member registered and the purchases where the member was the buyer, the treasurer will take their place
        db.execute("UPDATE purchases SET member_id = ? WHERE member_id = ?", treasurer_id, member_id)
        db.execute("UPDATE purchases SET buyer_id = ? WHERE buyer_id = ?", treasurer_id, member_id)
        db.execute("UPDATE earnings SET member_id = ? WHERE member_id = ?", treasurer_id, member_id)
        
        # delete member's data from members list and user accounts
        db.execute("DELETE FROM users WHERE member_id = ?", member_id)
        db.execute("DELETE FROM members WHERE id = ?", member_id)
    
    # go back to members list
    return redirect('/members')




@app.route("/board")
@login_required
def board():
    ''' Show the group's board'''
    
    # get the data of the board and display it
    board = db.execute('SELECT member_id, name, position FROM board JOIN members ON board.member_id = members.id')
    return render_template('board_list.html', board=board)


@app.route("/add_board", methods=["GET", "POST"])
@login_required
def add_board():
    ''' Add a new member to the board '''
    
    if request.method == 'POST':
        
        # get the member that joins the board and their position
        member_id = request.form.get('member')
        position = request.form.get('position')
        
        # ensure a member was picked
        if not member_id:
            flash('Choose a new board member')
            return redirect('/add_board')
        
        # ensure a position was typed in
        elif not position:
            flash('Type a board position')
            return redirect('/add_board')
        
        # insert new row into the database
        db.execute("INSERT INTO board (member_id, position) VALUES (?, ?)", member_id, position)
        
        # go back to the board list
        return redirect('/board')
    
    else:
        # get list of all members to be displayed in appropriate template
        members = db.execute("SELECT name, id FROM members WHERE id NOT IN (SELECT member_id FROM board JOIN members ON board.member_id = members.id)")
        return render_template('add_board.html', members=members)


@app.route("/delete_board", methods=["GET", "POST"])
@login_required
def delete_board():
    ''' Delete a board member '''
    
    if request.method == 'POST':
        # get the id of the board member to delete
        member_id = request.form.get("member_id")
        
        # delete member from database
        db.execute("DELETE FROM board WHERE member_id = ?", member_id)
    
    # return to board list
    return redirect('/board')


@app.route("/edit_board", methods=["GET", "POST"])
@login_required
def edit_board():
    ''' Loads page to edit board member '''
    
    if request.method == 'POST':
        # get id of member to edit
        member_id = request.form.get("member_id")
        
        # get their info to autofill values in html form
        person = db.execute("SELECT name, position, member_id FROM board JOIN members ON board.member_id = members.id WHERE member_id = ?", member_id)[0]
        
        # render template with values
        return render_template('edit_board.html', person=person)
   
    else:
        return redirect('/board')


@app.route("/editing_board", methods=["GET", "POST"])
@login_required
def editing_board():
    ''' Finalizes editing a board member '''
    
    
    if request.method == 'POST':
        
        # get the member's id and board position
        member_id = request.form.get("member_id")
        position = request.form.get('position')
        
        # ensure a position was submitted
        if not position:
            flash('Type a board position')
            person = db.execute("SELECT name, position, member_id FROM board JOIN members ON board.member_id = members.id WHERE member_id = ?", member_id)[0]
            return render_template('edit_board.html', person=person)
        
        # update database with new data
        db.execute("UPDATE board SET position = ? WHERE member_id = ?", position, member_id)
    
    # go back to board list   
    return redirect('/board')





    


''' HANDLING PURCHASES & EARNINGS '''

@app.route("/purchases")
@login_required
def purchases():
    ''' Display purchases'''
    
    # get all purchases from database
    purchases = db.execute("SELECT * FROM purchases")
    
    # loop through and format for display
    for purchase in purchases:
        
        # format date string from year, month, day data
        purchase['date'] = "{0}/{1}/{2}".format(str(purchase['month']).zfill(2), str(purchase['day']).zfill(2), str(purchase['year']))
        
        # get the names of the group members who registered the purchase and who was registered as the buyer
        buyer_name = db.execute('SELECT name FROM members WHERE id = ?', purchase['buyer_id'])
        registree_name = db.execute('SELECT name FROM members WHERE id = ?', purchase['member_id'])
        
        # format when there was no buyer registered
        if buyer_name:
            purchase['buyer'] = buyer_name[0]['name']
        else:
            purchase['buyer'] = '-'
        
        # format total in dollars
        purchase['total'] = "$ {:.2f}".format(purchase['total'])
    
    # render template with info
    return render_template('purchases.html', purchases=purchases)


@app.route("/add_purchase", methods=["GET", "POST"])
@login_required
def add_purchase():
    ''' Add new purchase '''
    
    if request.method == 'POST':
        
        # get data inputted
        date = request.form.get('date')
        total = request.form.get('total')
        vendor = request.form.get('vendor')
        
        # ensure date was submitted
        if not date:
            flash('Select a date')
            return redirect('/add_purchase')
        
        # ensure total was submitted
        elif not total:
            flash('Type in total')
            return redirect('/add_purchase')
        
        # ensure vendor was submitted
        elif not vendor:
            flash('Type in vendor')
            return redirect('/add_purchase')
        
        # get buyer_id
        buyer_id = request.form.get('buyer_id')
        
        # get description and format if empty
        description = request.form.get('desc')
        if not description:
            description = '-'
        
        # split html date value into year, month and day
        date_values = str(date).split('-')
        year = date_values[0]
        month = date_values[1]
        day = date_values[2]
        
        # get the member_id of the active user to be used as the member who registered the purchase
        user_id = session["user_id"]
        member_id = db.execute('SELECT member_id FROM users WHERE id = ?', user_id)[0]['member_id']
        
        # insert new row into database
        db.execute('INSERT INTO purchases (member_id, total, place, buyer_id, year, month, day, description) VALUES (?, ?, ?, ?, ?, ?, ?, ?)', member_id, total, vendor, buyer_id, year, month, day, description)

        # go back to purchases
        return redirect('/purchases')
    
    else:
        # send the list of members for dropdowns
        members = db.execute("SELECT name, id FROM members")
        return render_template('add_purchase.html', members=members)


@app.route("/edit_purchase", methods=["GET", "POST"])
@login_required
def edit_purchase():
    ''' Loads page to edit a purchase '''
    
    if request.method == 'POST':
        
        # get the id of the purchase to edit
        purchase_id = request.form.get("id")
        
        # get the purchase info and format date
        purchase = db.execute("SELECT id, total, place, buyer_id, description, year, month, day FROM purchases WHERE id = ?", purchase_id)[0] 
        purchase['date'] = "{0}-{1}-{2}".format(purchase['year'], str(purchase['month']).zfill(2), str(purchase['day']).zfill(2))
        
        # get list of members for dropdown menu
        members = db.execute("SELECT name, id FROM members")
        
        # format page with values for html form
        return render_template('edit_purchase.html', purchase=purchase, members=members)
    
    return redirect('/')


@app.route("/editing_purchase", methods=["GET", "POST"])
@login_required
def editing_purchase():
    ''' Finalizes editing a purchase '''
    
    
    if request.method == 'POST':
        
        # template to render if inputs are incorrect
        template = 'edit_purchase.html'
        
        # get data from form
        purchase_id = request.form.get('purchase_id')
        date = request.form.get('date')
        total = request.form.get('total')
        vendor = request.form.get('vendor')
        
        # ensure date was submitted
        if not date:
            flash('Select a date')
            purchase = db.execute("SELECT id, total, place, buyer_id, description, year, month, day FROM purchases WHERE id = ?", purchase_id)[0] 
            purchase['date'] = "{0}-{1}-{2}".format(purchase['year'], str(purchase['month']).zfill(2), str(purchase['day']).zfill(2))
            members = db.execute("SELECT name, id FROM members")
            return render_template(template, purchase=purchase, members=members)
        
        # ensure total was submitted
        elif not total:
            flash('Type in total')
            purchase = db.execute("SELECT id, total, place, buyer_id, description, year, month, day FROM purchases WHERE id = ?", purchase_id)[0] 
            purchase['date'] = "{0}-{1}-{2}".format(purchase['year'], str(purchase['month']).zfill(2), str(purchase['day']).zfill(2))
            members = db.execute("SELECT name, id FROM members")
            return render_template(template, purchase=purchase, members=members)
        
        # ensure vendor was submitted
        elif not vendor:
            flash('Type in vendor')
            purchase = db.execute("SELECT id, total, place, buyer_id, description, year, month, day FROM purchases WHERE id = ?", purchase_id)[0] 
            purchase['date'] = "{0}-{1}-{2}".format(purchase['year'], str(purchase['month']).zfill(2), str(purchase['day']).zfill(2))
            members = db.execute("SELECT name, id FROM members")
            return render_template(template, purchase=purchase, members=members)
        
        # get buyer_id
        buyer_id = request.form.get('buyer_id')
        
        # get and format description
        description = request.form.get('desc')
        if not description:
            description = '-'
        
        # format date from html
        date_values = str(date).split('-')
        year = date_values[0]
        month = date_values[1]
        day = date_values[2]
        
        # get current user to be saved as person registering purchase
        user_id = session["user_id"]
        member_id = db.execute('SELECT member_id FROM users WHERE id = ?', user_id)[0]['member_id']
        
        # update database
        db.execute("UPDATE purchases SET member_id = ?, total = ?, place = ?, buyer_id = ?, description = ?, year = ?, month = ?, day = ? WHERE id = ?", member_id, total, vendor, buyer_id, description, year, month, day, purchase_id)
        
    return redirect('/purchases')
    

@app.route("/delete_purchase", methods=["GET", "POST"])
@login_required
def delete_purchase():
    ''' Deletes a purchase '''
    
    if request.method == 'POST':
        
        # get id of purchase to be deleted
        purchase_id = request.form.get('id')
        
        # remove row from database
        db.execute("DELETE FROM purchases WHERE id = ?", purchase_id)
       
    # go back to purchases list 
    return redirect('/purchases')





@app.route("/earnings")
@login_required
def earnings():
    ''' Display earnings '''
    
    # get all the earnings from database
    earnings = db.execute("SELECT * FROM earnings")
    
    # format each earning
    for earning in earnings:
        
        # format date as string
        earning['date'] = "{0}/{1}/{2}".format(str(earning['month']).zfill(2), str(earning['day']).zfill(2), earning['year'])
        
        # get name of member who registed purchase
        earning['registree'] = db.execute('SELECT name FROM members WHERE id = ?', earning['member_id'])[0]['name']
        
        # format total as dollars
        earning['total'] = "$ {:.2f}".format(earning['total'])
    
    # render template with correct info  
    return render_template("earnings.html", earnings=earnings)


@app.route("/add_earning", methods=["GET", "POST"])
@login_required
def add_earning():
    ''' Add new earning '''
    
    if request.method == 'POST':
        
        # get data from form
        date = request.form.get('date')
        total = request.form.get('total')
        source = request.form.get('source')
        
        # ensure date was submitted
        if not date:
            flash('Select a date')
            return redirect('/add_earning')
        
        # ensure total was submitted
        elif not total:
            flash('Type in total')
            return redirect('/add_earning')
        
        # ensure source was submitted
        elif not source:
            flash('Type in source')
            return redirect('/add_earning')
        
        # format description
        description = request.form.get('desc')
        if not description:
            description = '-'
        
        # format date from html values
        date_values = str(date).split('-')
        year = date_values[0]
        month = date_values[1]
        day = date_values[2]
        
        # get current member to be used as member who registered the earning
        user_id = session["user_id"]
        member_id = db.execute('SELECT member_id FROM users WHERE id = ?', user_id)[0]['member_id']
        
        # insert new row into database
        db.execute('INSERT INTO earnings (member_id, total, source, year, month, day, notes) VALUES (?, ?, ?, ?, ?, ?, ?)', member_id, total, source, year, month, day, description)

        # go back to earnings page
        return redirect('/earnings')
    
    else:
        return render_template('add_earning.html')


@app.route("/edit_earning", methods=["GET", "POST"])
@login_required
def edit_earning():
    ''' Loads page to edit earning '''
    
    if request.method == 'POST':
        
        # get id of earning to be editted
        earning_id = request.form.get('earning_id')
        
        # get the earning info to be passed as values for html form
        earning_info = db.execute("SELECT id, year, month, day, total, source, notes FROM earnings WHERE id = ?", earning_id)[0]
        earning_info['date'] = "{0}-{1}-{2}".format(earning_info['year'], str(earning_info['month']).zfill(2), str(earning_info['day']).zfill(2))
        
        # render the template with the values
        return render_template('edit_earning.html', earning=earning_info)
        
    return redirect('/')

@app.route("/editing_earning", methods=["GET", "POST"])
@login_required
def editing_earning():
    ''' Finalizes editing an earning '''
    
    if request.method == 'POST':
        
        # template to render if inputs are incorrect
        template = 'edit_earning.html'
        
        # get data from form
        earning_id = request.form.get('earning_id')
        date = request.form.get('date')
        total = request.form.get('total')
        source = request.form.get('source')
        
        # ensure date was submitted
        if not date:
            flash('Select a date')
            earning_info = db.execute("SELECT id, year, month, day, total, source, notes FROM earnings WHERE id = ?", earning_id)[0]
            earning_info['date'] = "{0}-{1}-{2}".format(earning_info['year'], str(earning_info['month']).zfill(2), str(earning_info['day']).zfill(2))
            return render_template(template, earning=earning_info)
        
        # ensure total was submitted
        elif not total:
            flash('Type in total')
            earning_info = db.execute("SELECT id, year, month, day, total, source, notes FROM earnings WHERE id = ?", earning_id)[0]
            earning_info['date'] = "{0}-{1}-{2}".format(earning_info['year'], str(earning_info['month']).zfill(2), str(earning_info['day']).zfill(2))
            return render_template(template, earning=earning_info)
        
        # ensure source was submitted
        elif not source:
            flash('Type in source')
            earning_info = db.execute("SELECT id, year, month, day, total, source, notes FROM earnings WHERE id = ?", earning_id)[0]
            earning_info['date'] = "{0}-{1}-{2}".format(earning_info['year'], str(earning_info['month']).zfill(2), str(earning_info['day']).zfill(2))
            return render_template(template, earning=earning_info)
        
        # format description
        description = request.form.get('desc')
        if not description:
            description = '-'
        
        # format date
        date_values = str(date).split('-')
        year = date_values[0]
        month = date_values[1]
        day = date_values[2]
        
        # get current member to be used as member who registered the earning
        user_id = session["user_id"]
        member_id = db.execute('SELECT member_id FROM users WHERE id = ?', user_id)[0]['member_id']
        
        # update row in database
        db.execute('UPDATE earnings SET member_id = ?, total = ?, source = ?, year = ?, month = ?, day = ?, notes = ? WHERE id = ?', member_id, total, source, year, month, day, description, earning_id)

    # go back to list of earnings
    return redirect('/earnings')


@app.route("/delete_earning", methods=["GET", "POST"])
@login_required
def delete_earning():
    ''' Delete earning '''
    
    if request.method == 'POST':
        
        # get id of earning to be deleted
        earning_id = request.form.get('earning_id')
        
        # delete from the database
        db.execute("DELETE FROM earnings WHERE id = ?", earning_id)
    
    # go back to earnings list
    return redirect('/earnings')





''' FINANCE ANALYTICS '''

@app.route("/finance_analytics", methods=["GET", "POST"])
@login_required
def finance_analytics():
    ''' Renders a plot for purchases and a plot for earnings'''
    
    # create earnings and purchases plots with a function
    earnings_script, earnings_plot = monthly_totals_graph('e')
    purchases_script, purchases_plot = monthly_totals_graph('p')
    
    # render plots in appropriate template
    return render_template('finance_analytics.html', earnings_script=earnings_script, earnings_plot=earnings_plot , purchases_script=purchases_script, purchases_plot=purchases_plot)


@app.route('/return_excel/', methods=["GET", "POST"])
@login_required
def return_excel():
    ''' Calls function to update excel file and returns file as an attachment '''
    
    if request.method == 'POST':
        
        # get the start and end dates selected by the user
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        
        # ensure both a start date and an end date are selected or that neither is selected
        if start_date and not end_date:
            flash('Select and end date if you are selecting a start date')
            return redirect('/finance_analytics')
        elif end_date and not start_date:
            flash('Select a start date if you are selecting an end date')
            return redirect('/finance_analytics')
        
        # call function to update excel file
        update_finances_excel(start_date, end_date)
        
        # send excel as an attachment
        try:
            return send_file(path, as_attachment=True)
        except Exception as e:
            return str(e)
    else:
        return redirect('/finance_analytics')
 

def update_finances_excel(start_date, end_date):
    
    '''
    Saves earnings and purchases from (start_date) to (end_date) in Excel file
        If no dates are provided, all purchases and earnings are saved
    '''
    
    # open a workbook and a worksheet
    workbook = load_workbook(path)
    sheet =  workbook.active
    
    #https://www.folkstalk.com/2022/10/how-to-delete-every-row-in-excel-using-openpyxl-with-code-examples.html
    # delete everything in the worksheet
    sheet.delete_rows(1, sheet.max_row+1)
    
    # append the headers to the sheet
    headers = ['Date', 'Total', 'Vendor / Source', 'Registered by', 'Description']
    sheet.append(headers)
    
    
    # get the earnings and purchases data
    if not start_date and not end_date:
        # get all if no dates were provided
        purchases = db.execute('SELECT year, month, day, name, total, place, description FROM purchases JOIN members ON purchases.member_id = members.id')
        earnings = db.execute('SELECT year, month, day, name, total, source, notes FROM earnings JOIN members ON earnings.member_id = members.id')
    
    else:
        # format the dates to use them when looking up purchases and earnings
        sdate = str(start_date).split('-')
        edate = str(end_date).split('-')
        
        # get appropriate ranges if dates were provided
        purchases = db.execute('SELECT year, month, day, name, total, place, description FROM purchases JOIN members ON purchases.member_id = members.id WHERE (year BETWEEN ? AND ?) AND (month BETWEEN ? AND ?) AND (day BETWEEN ? AND ?)', sdate[0], edate[0], sdate[1], edate[1], sdate[2], edate[2])
        earnings = db.execute('SELECT year, month, day, name, total, source, notes FROM earnings JOIN members ON earnings.member_id = members.id WHERE (year BETWEEN ? AND ?) AND (month BETWEEN ? AND ?) AND (day BETWEEN ? AND ?)', sdate[0], edate[0], sdate[1], edate[1], sdate[2], edate[2])
    
    # space to save formated version of the data
    purchases_formatted = []
    earnings_formatted = []
    
    # tutorialspoint.com/How-to-convert-Python-Dictionary-to-a-list
    # format each purchase
    for purchase in purchases:
        # all purchases start with "-"
        purchase['total'] = "-{0}".format(purchase['total'])
        purchase['date'] = "{0}/{1}/{2}".format(purchase['month'], purchase['day'], purchase['year'])
        purchases_formatted.append([purchase['date'], purchase['total'], purchase['place'], purchase['name'], purchase['description']])
    
    # format each earning
    for earning in earnings:
        # all earnings start with "+"
        earning['total'] = "+{0}".format(earning['total'])
        earning['date'] = "{0}/{1}/{2}".format(earning['month'], earning['day'], earning['year'])
        earnings_formatted.append([earning['date'], earning['total'], earning['source'], earning['name'], earning['notes']])
    
    # append formatted purchases to Excel
    for purchase in purchases_formatted:
        sheet.append(purchase)
    
    # append formatted earnings to Excel
    for earning in earnings_formatted:
        sheet.append(earning)
    
    # save the Excel file
    workbook.save(path)


def monthly_totals_graph(type):
    ''' Makes a graph of monthly totals for purchases or earnings. Type:  e = earnings, p = purchases. Returns <script> and <div> elements to be included in HTML'''
    
    # dictionary to store the totals for each month
    monthly_totals = {}
    
    # check if the graph should be made for earnings or purchases
    if type == 'e':
        
        # get the earnings
        earnings = db.execute('SELECT year, month, total FROM earnings')
        
        # loop through each earning to sum up its value
        for earning in earnings:
            # change the month number into its name
            month = calendar.month_name[earning['month']]
            year = earning['year']
            
            # get total
            total = earning['total']
            
            # create a tupple with the formatted month and the year
            key = (month, year)
            
            # use the key tuple to update the appropriate key-value pair or create a new pair if it doesn't exist
            monthly_totals[key] = monthly_totals.get(key, 0) + total
        
        # set the color of the graph
        color = 'green'
        
        # set legend of graph
        legend = "Earnings totals"
    
    else:
        # get the purchases
        purchases = db.execute('SELECT year, month, total FROM purchases')
        
        # loop through each purchase to sum up its value
        for purchase in purchases:
            # change the month number into its name
            month = calendar.month_name[purchase['month']]
            year = purchase['year']
            
            # get total
            total = purchase['total']
        
            # create a tupple with the formatted month and the year
            key = (month, year)
            
            # use the key tuple to update the appropriate key-value pair or create a new pair if it doesn't exist
            monthly_totals[key] = monthly_totals.get(key, 0) + total
        
        # set the color of the graph
        color = 'blue'
        
        # set legend of graph
        legend = "Earnings totals"
    
    
    # space to store data for graphing
    x = []
    y = []

    # iterates over key-value pairs in the monthly_totals dictionary
    for (year, month), total in monthly_totals.items():
        # format month year values and appends them to x list
        x.append(f'{month} {year}')
        
        # appends totals to y list
        y.append(total)
    
    
    # creates a zip object to pair up appropriate x, y pairs
    xy_pairs = zip(x, y)
    
    # sort the zip in descending order based on month-year values (i.e. set 2022 september before 2022 november)
    xy_pairs = sorted(xy_pairs, key=lambda pair: pair[0], reverse=True)
    
    # unzip the sorted values and send them back to the correct list
    x, y = zip(*xy_pairs)
    
    # create a range for the x-axis based on the values from the x list
    # this is needed to have non numerical values in the x axis of the Bokeh plot
    x_range = FactorRange(factors=x)
    
    # create a new Bokeh plot with a title and axis labels
    p = figure(title=None, x_axis_label="month", y_axis_label="total", x_range=x_range, min_border_bottom=0)

    # add a bars with colors and a legend
    p.vbar(x=x, top=y, legend_label=legend, width=0.2, bottom=0, color=color)

    # show the results
    components(p)
    
    return components(p)